from openpyxl import Workbook, load_workbook


# 打开现有的 Excel 文件或创建新的 Excel 文件

def save(filename, acc, model_name, ds_name):
    if filename is None:
        filename = r'./results/accuracy.xlsx'
    try:
        workbook = load_workbook(filename=filename)
    except FileNotFoundError:
        workbook = Workbook()

    # 选择或创建名为“model_name”的工作表
    try:
        sheet = workbook[model_name]
    except KeyError:
        # sheet = workbook.active
        # sheet.title = model_name
        sheet = workbook.create_sheet(title=model_name)
        sheet.cell(row=1, column=1, value='dataset')

    last_row = sheet.max_row

    # save dataset and accuracy
    sheet.cell(row=last_row + 1, column=1, value=ds_name)
    for i in range(len(acc)):
        sheet.cell(row=last_row + 1, column=i + 2, value=acc[i])
    # 保存 Excel 文件
    workbook.save(filename=filename)

if __name__ == '__main__':
    save([0.12, 0.12, 0.32], 'ts2vec', 'Adiac')
